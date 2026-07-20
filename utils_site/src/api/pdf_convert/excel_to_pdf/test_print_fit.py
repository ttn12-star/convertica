"""Runnable check for the Excel->PDF print-fit preprocessing.

Pure openpyxl assertions (no LibreOffice) so it is safe in CI. Verifies the
branch that fixes wide tables spilling across PDF pages ("table shifted").

Run standalone:  python test_print_fit.py
Or via pytest:   pytest test_print_fit.py
"""

import os
import tempfile

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties
from src.api.pdf_convert.excel_to_pdf.utils import _apply_print_fit


def _make_xlsx(path, cols, col_width=None, preconfigured=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    for j in range(1, cols + 1):
        ws.cell(1, j, f"C{j}")
        if col_width:
            ws.column_dimensions[get_column_letter(j)].width = col_width
    if preconfigured:  # author already set a manual print scale
        ws.page_setup.scale = 60
    wb.save(path)


def test_wide_table_fits_and_goes_landscape():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "wide.xlsx")
        _make_xlsx(p, cols=25, col_width=14)
        _apply_print_fit(p, {})
        ws = openpyxl.load_workbook(p).active
        assert ws.page_setup.fitToWidth == 1
        assert ws.page_setup.fitToHeight == 0
        assert ws.sheet_properties.pageSetUpPr.fitToPage is True
        assert ws.page_setup.orientation == "landscape"


def test_narrow_table_fits_but_stays_portrait():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "narrow.xlsx")
        _make_xlsx(p, cols=4)  # default width -> well under threshold
        _apply_print_fit(p, {})
        ws = openpyxl.load_workbook(p).active
        assert ws.page_setup.fitToWidth == 1
        # not forced into landscape
        assert ws.page_setup.orientation != "landscape"


def test_author_print_scale_is_respected():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "authored.xlsx")
        _make_xlsx(p, cols=25, col_width=14, preconfigured=True)
        _apply_print_fit(p, {})
        ws = openpyxl.load_workbook(p).active
        # We must not clobber a deliberate manual scale.
        assert ws.page_setup.scale == 60
        assert ws.page_setup.fitToWidth != 1


def test_non_xlsx_is_noop():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "legacy.xls")
        with open(p, "wb") as f:
            f.write(b"\xd0\xcf\x11\xe0not-a-real-xls")
        before = os.path.getsize(p)
        _apply_print_fit(p, {})  # must not raise
        assert os.path.getsize(p) == before


def test_forced_portrait_overrides_wide_heuristic():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "wide.xlsx")
        _make_xlsx(p, cols=25, col_width=14)  # would auto-go landscape
        _apply_print_fit(p, {}, orientation="portrait")
        ws = openpyxl.load_workbook(p).active
        assert ws.page_setup.orientation == "portrait"
        assert ws.page_setup.fitToWidth == 1  # still fits width


def test_actual_mode_leaves_size_untouched():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "wide.xlsx")
        _make_xlsx(p, cols=25, col_width=14)
        _apply_print_fit(p, {}, fit_mode="actual")
        ws = openpyxl.load_workbook(p).active
        # No fit-to-width forced -> table keeps native size (may span pages).
        assert ws.page_setup.fitToWidth != 1


def test_user_choice_overrides_author_scale():
    with tempfile.TemporaryDirectory() as d:
        p = os.path.join(d, "authored.xlsx")
        _make_xlsx(p, cols=25, col_width=14, preconfigured=True)
        # Explicit user pick must win over the author's manual scale.
        _apply_print_fit(p, {}, orientation="landscape")
        ws = openpyxl.load_workbook(p).active
        assert ws.page_setup.orientation == "landscape"


if __name__ == "__main__":
    test_wide_table_fits_and_goes_landscape()
    test_narrow_table_fits_but_stays_portrait()
    test_author_print_scale_is_respected()
    test_non_xlsx_is_noop()
    test_forced_portrait_overrides_wide_heuristic()
    test_actual_mode_leaves_size_untouched()
    test_user_choice_overrides_author_scale()
    print("OK: all print-fit checks passed")
